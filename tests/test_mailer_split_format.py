"""Split & Send must hand each recipient a file that looks like the report.

Field report: a formatted MoM workbook came out the other side as a bare grid —
no header band, default column widths, money as raw numbers. The split read
values only and wrote a fresh sheet, so every style was discarded by design.

Also pinned here: rows that carry no address but DO name their owner (the banner
heading each person's block) reach that person, while rows owned by nobody stay
unsent.
"""

from __future__ import annotations

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.mailer_split import split_by_email


def _source(tmp_path, *, owner_header="Sales Manager"):
    wb = Workbook()
    ws = wb.active
    ws.append(["Agent", owner_header, f"{owner_header} Email", "Net"])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF", size=12)
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.alignment = Alignment(horizontal="center")
    ws.append(["KAM: Abdul Aziz", "Abdul Aziz", None, None])       # banner, no email
    ws.append(["First Trip", "Abdul Aziz", "aziz@x.com", 228031544])
    ws.append(["Sky High", "Abdul Aziz", "aziz@x.com", 1234])
    ws.append(["Bravo", "Rahim", "rahim@x.com", 5678])
    ws.append(["Orphan Ltd", "(no KAM assigned)", None, 99])       # owned by nobody
    for r in range(3, 7):
        ws.cell(r, 4).number_format = "#,##0_);[Red](#,##0)"
    ws.column_dimensions["A"].width = 45.4
    ws.column_dimensions["B"].width = 40.2
    ws.row_dimensions[1].height = 26
    p = tmp_path / "MoM.xlsx"
    wb.save(p)
    return p


def _split(tmp_path, **kw):
    src = _source(tmp_path)
    out = tmp_path / "out"
    return src, split_by_email(src, out, email_column="Sales Manager Email", **kw)


def test_header_band_survives(tmp_path):
    _src, r = _split(tmp_path)
    ws = load_workbook(next(g.path for g in r.groups if g.email == "aziz@x.com")).active
    h = ws.cell(1, 1)
    assert h.fill.fgColor.rgb.endswith("1F3864")
    assert h.font.b and h.font.sz == 12 and h.font.color.rgb.endswith("FFFFFF")


def test_column_widths_and_row_height_survive(tmp_path):
    _src, r = _split(tmp_path)
    ws = load_workbook(r.groups[0].path).active
    assert round(ws.column_dimensions["A"].width, 1) == 45.4
    assert round(ws.column_dimensions["B"].width, 1) == 40.2
    assert ws.row_dimensions[1].height == 26


def test_number_format_survives_so_money_is_not_a_raw_number(tmp_path):
    _src, r = _split(tmp_path)
    ws = load_workbook(next(g.path for g in r.groups if g.email == "aziz@x.com")).active
    vals = {ws.cell(row, 1).value: row for row in range(2, ws.max_row + 1)}
    assert ws.cell(vals["First Trip"], 4).number_format == "#,##0_);[Red](#,##0)"
    assert ws.cell(vals["First Trip"], 4).value == 228031544


def test_header_row_is_frozen(tmp_path):
    _src, r = _split(tmp_path)
    assert load_workbook(r.groups[0].path).active.freeze_panes == "A2"


def test_banner_row_reaches_its_owner(tmp_path):
    """The heading of the block being sent used to be dropped from every copy."""
    _src, r = _split(tmp_path)
    ws = load_workbook(next(g.path for g in r.groups if g.email == "aziz@x.com")).active
    first_col = [ws.cell(row, 1).value for row in range(1, ws.max_row + 1)]
    assert "KAM: Abdul Aziz" in first_col
    assert any("named an owner" in w for w in r.warnings)


def test_a_row_owned_by_nobody_is_not_sent_to_anybody(tmp_path):
    """'(no KAM assigned)' resolves to no one — it must stay unmatched, not be
    guessed onto the nearest recipient."""
    _src, r = _split(tmp_path)
    for g in r.groups:
        col = [c.value for row in load_workbook(g.path).active.iter_rows()
               for c in row]
        assert "Orphan Ltd" not in col
    assert r.unmatched_path and len(r.unmatched_rows) == 1


def test_other_recipients_rows_never_leak(tmp_path):
    _src, r = _split(tmp_path)
    ws = load_workbook(next(g.path for g in r.groups if g.email == "aziz@x.com")).active
    col = [ws.cell(row, 1).value for row in range(1, ws.max_row + 1)]
    assert "Bravo" not in col and "Sky High" in col


def test_owner_column_can_be_named_explicitly(tmp_path):
    src = _source(tmp_path)
    r = split_by_email(src, tmp_path / "o2", email_column="Sales Manager Email",
                       owner_column="Sales Manager")
    assert any("named an owner" in w for w in r.warnings)


def test_no_owner_column_means_no_recovery(tmp_path):
    """With nothing to resolve from, a blank address must still be unmatched."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Agent", "Mail"])
    ws.append(["Banner", None])
    ws.append(["Real", "a@x.com"])
    src = tmp_path / "plain.xlsx"
    wb.save(src)
    r = split_by_email(src, tmp_path / "o3", email_column="Mail")
    assert len(r.unmatched_rows) == 1 and len(r.groups) == 1
