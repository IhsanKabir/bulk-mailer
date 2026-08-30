"""Split one main sheet into per-recipient workbooks for the Bulk Mailer.

The user keeps ONE Excel where every data row carries an email address in some
column. `split_by_email` groups the rows by each address found there
(case-insensitive; a cell may list several addresses separated by ; , or |,
in which case the row goes to EACH), writes one workbook per address into a
chosen folder, and parks blank/invalid-address rows in `_UNMATCHED_ROWS.xlsx`
so nothing is ever silently dropped. `build_mail_rows` then wraps each split
file as a ready-to-send `mailer_io.MailRow` — the exact shape the existing
Bulk Mailer preview/run pipeline consumes — stamping a GUI-level CC/BCC on
every message and exposing {email}/{name}/{rows}/{file} template fields.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook

from .mailer_io import MailRow

log = logging.getLogger(__name__)

_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
_ADDR_SPLIT_RE = re.compile(r"[;,|]")
_UNSAFE_FILENAME_RE = re.compile(r'[\\/:*?"<>|\s]+')
UNMATCHED_FILENAME = "_UNMATCHED_ROWS.xlsx"
_MAX_STEM_LEN = 100


@dataclass(frozen=True)
class SplitGroup:
    """One recipient's slice: their address, how many rows, and the file."""

    email: str
    row_count: int
    path: str


@dataclass(frozen=True)
class SplitResult:
    """Outcome of one split run. `unmatched_rows` are 1-based data-row numbers
    whose email cell was blank/invalid — written to `unmatched_path`."""

    source: str
    email_column: str
    groups: tuple[SplitGroup, ...]
    unmatched_rows: tuple[int, ...]
    unmatched_path: str | None
    warnings: tuple[str, ...]


def read_headers(path: str | Path, sheet_name: str | None = None) -> list[str]:
    """First-row header labels of the sheet (for the email-column picker)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active
        for row in ws.iter_rows(max_row=1, values_only=True):
            return ["" if c is None else str(c).strip() for c in row]
        return []
    finally:
        wb.close()


def list_sheet_names(path: str | Path) -> list[str]:
    """Sheet names of a workbook (feeds the sheet picker)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _addresses_in(cell: str) -> list[str]:
    """Valid, lowercased addresses in one cell (may hold several)."""
    return [a.strip().lower() for a in _ADDR_SPLIT_RE.split(cell)
            if a.strip() and _EMAIL_RE.match(a.strip())]


def _safe_stem(email: str, used: set[str]) -> str:
    """Filesystem-safe unique file stem for an address ('@' is legal on Windows)."""
    stem = _UNSAFE_FILENAME_RE.sub("_", email)[:_MAX_STEM_LEN] or "recipient"
    base, n = stem, 2
    while stem.lower() in used:
        stem = f"{base}_{n}"
        n += 1
    used.add(stem.lower())
    return stem


def _copy_cell(src, dst, cache: dict) -> None:
    """Value AND presentation. `_style` carries font, fill, border, alignment and
    number format in one object, so a date stays a date and a header stays styled."""
    dst.value = src.value
    if not src.has_style:
        return
    # A cell's _style is an array of INDEXES into its own workbook's font/fill/
    # alignment tables, so it cannot be copied across workbooks (openpyxl raises
    # IndexError on save). The style objects have to be assigned instead — but
    # each assignment searches and appends to the destination's style tables, and
    # at five per cell over half a million cells that dominated the runtime.
    #
    # So each DISTINCT source style is registered in the destination exactly once
    # (on a scratch cell), its resulting index array captured, and every later
    # cell with that style gets the array in a single assignment.
    key = src.style_id
    arr = cache.get(key)
    if arr is None:
        probe = cache["__probe__"]
        probe.font = copy(src.font)
        probe.fill = copy(src.fill)
        probe.border = copy(src.border)
        probe.alignment = copy(src.alignment)
        probe.protection = copy(src.protection)
        probe.number_format = src.number_format
        arr = cache[key] = copy(probe._style)
    dst._style = copy(arr)


def _write_rows(path: Path, ws_src, header_row: int, source_rows: list[int],
                max_col: int) -> None:
    """One recipient's workbook, looking like the report it came from.

    The split used to read values only and write a bare sheet, so every recipient
    got an unformatted grid: no header band, default column widths, dates and
    money as raw numbers. Rows are copied cell-by-cell WITH their style, along
    with the column widths and row heights, so the file the recipient opens looks
    like the one the analyst built.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Rows"
    # the scratch cell used to register styles lives on a sheet we drop before save
    scratch = wb.create_sheet("__styles__")
    style_cache: dict = {"__probe__": scratch.cell(row=1, column=1)}
    for key, dim in ws_src.column_dimensions.items():
        if dim.width:
            ws.column_dimensions[key].width = dim.width
            ws.column_dimensions[key].hidden = dim.hidden
    out_r = 1
    for src_r in [header_row] + list(source_rows):
        for c in range(1, max_col + 1):
            _copy_cell(ws_src.cell(src_r, c), ws.cell(out_r, c), style_cache)
        height = ws_src.row_dimensions[src_r].height
        if height:
            ws.row_dimensions[out_r].height = height
        out_r += 1
    # the source's own freeze point cannot survive filtering; keep the header
    ws.freeze_panes = "A2"
    wb.remove(scratch)
    wb.save(path)


def split_by_email(
    path: str | Path,
    out_dir: str | Path,
    *,
    email_column: str,
    sheet_name: str | None = None,
    owner_column: str | None = None,
) -> SplitResult:
    """Split the sheet into one workbook per email address in `email_column`.

    Each recipient's file keeps the source's FORMATTING — header band, column
    widths, number formats — because rows are copied from the sheet with their
    styles rather than re-written as bare values.

    Raises ValueError when the column is missing (fail fast at the boundary);
    data problems (blank/invalid addresses, empty sheet) come back as warnings
    + the unmatched file instead of exceptions, so one bad row never kills a run.
    """
    path = Path(path)
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)

    # NOT read_only: styles are needed to carry the source formatting across
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active
        max_col = ws.max_column
        header_row = 1
        rows_iter = ws.iter_rows()
        try:
            headers = ["" if c.value is None else str(c.value).strip()
                       for c in next(rows_iter)]
        except StopIteration:
            return SplitResult(str(path), email_column, (), (), None,
                               ("Sheet is empty — nothing to split.",))
        wanted = email_column.strip().lower()
        try:
            email_i = [h.lower() for h in headers].index(wanted)
        except ValueError:
            raise ValueError(
                f"Email column {email_column!r} not found. "
                f"Sheet headers: {', '.join(h for h in headers if h)}") from None

        # An OWNER column lets a row without an address still find its recipient.
        # A banded report puts a banner above each person's block ("KAM: <name>")
        # and only fills the address on the rows beneath it, so the banner — the
        # heading of the very block being sent — used to be dropped from everyone's
        # copy. When the owner is named and that name's address appears anywhere in
        # the sheet, the row goes to them. A row whose owner is a placeholder like
        # "(no KAM assigned)" resolves to nobody and stays unmatched, which is
        # correct: there is no one to send it to.
        owner_i = None
        if owner_column is None and email_column.strip().lower().endswith(" email"):
            owner_column = email_column.strip()[: -len(" email")]
        if owner_column:
            try:
                owner_i = [h.lower() for h in headers].index(owner_column.strip().lower())
            except ValueError:
                owner_i = None

        # keep the SOURCE ROW NUMBER, not just the values — the writer copies each
        # row, with its styling, from the sheet it came from
        scanned: list[tuple[int, int, str, str]] = []       # data,sheet,addr_cell,owner
        data_row = 0
        for cells in rows_iter:
            if not cells:
                continue
            if not any(str(c.value).strip() for c in cells if c.value is not None):
                continue                                    # trailing blank rows
            data_row += 1
            cell = "" if (email_i >= len(cells) or cells[email_i].value is None) \
                else str(cells[email_i].value).strip()
            owner = ""
            if owner_i is not None and owner_i < len(cells) \
                    and cells[owner_i].value is not None:
                owner = str(cells[owner_i].value).strip()
            scanned.append((data_row, cells[0].row, cell, owner))

        owner_to_address: dict[str, str] = {}
        for _d, _s, cell, owner in scanned:
            addrs = _addresses_in(cell)
            if owner and addrs:
                owner_to_address.setdefault(owner.lower(), addrs[0])

        by_email: dict[str, list[int]] = {}
        unmatched: list[tuple[int, int]] = []               # (data_row, sheet_row)
        recovered = 0
        for d_row, sheet_row, cell, owner in scanned:
            addresses = _addresses_in(cell)
            if not addresses and owner:
                inherited = owner_to_address.get(owner.lower())
                if inherited:
                    addresses = [inherited]
                    recovered += 1
            if not addresses:
                unmatched.append((d_row, sheet_row))
                continue
            for addr in addresses:                          # shared row -> each address
                by_email.setdefault(addr, []).append(sheet_row)

        warnings: list[str] = []
        if recovered:
            warnings.append(
                f"{recovered} row(s) had no address but named an owner in "
                f"{owner_column!r} — sent to that person.")
        if not by_email and not unmatched:
            warnings.append("Sheet is empty — nothing to split.")

        used: set[str] = set()
        groups: list[SplitGroup] = []
        for addr in sorted(by_email):
            gpath = out / f"{_safe_stem(addr, used)}.xlsx"
            _write_rows(gpath, ws, header_row, by_email[addr], max_col)
            groups.append(SplitGroup(email=addr, row_count=len(by_email[addr]),
                                     path=str(gpath)))

        unmatched_path: str | None = None
        if unmatched:
            unmatched_path = str(out / UNMATCHED_FILENAME)
            _write_rows(Path(unmatched_path), ws, header_row,
                        [r for _, r in unmatched], max_col)
            warnings.append(
                f"{len(unmatched)} row(s) had a blank/invalid address in "
                f"{email_column!r} — parked in {UNMATCHED_FILENAME}, NOT sent.")

        log.info("split %s: %d recipients, %d unmatched rows",
                 path.name, len(groups), len(unmatched))
        return SplitResult(
            source=str(path), email_column=email_column, groups=tuple(groups),
            unmatched_rows=tuple(i for i, _ in unmatched),
            unmatched_path=unmatched_path, warnings=tuple(warnings))
    finally:
        wb.close()


def build_mail_rows(result: SplitResult, *, cc: str = "", bcc: str = "") -> list[MailRow]:
    """Convert split groups into MailRows the existing mailer pipeline consumes.

    `cc`/`bcc` (the GUI's global fields) are stamped on EVERY message. Template
    fields per row: {email}, {name} (address local part), {rows}, {file}.
    """
    rows: list[MailRow] = []
    for i, g in enumerate(result.groups, start=1):
        p = Path(g.path)
        issues: tuple[str, ...] = () if p.is_file() else (f"file not found: {p.name}",)
        local = g.email.split("@", 1)[0]
        fields = {
            "email": g.email, "name": local, "Name": local,
            "rows": str(g.row_count), "file": p.name,
        }
        rows.append(MailRow(
            row_index=i, email=g.email, name=local, attachments=(p,),
            cc=cc.strip(), bcc=bcc.strip(), fields=fields, issues=issues))
    return rows
